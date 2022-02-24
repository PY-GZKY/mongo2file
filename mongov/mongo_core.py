# -*- coding:utf-8 -*-
import codecs
import csv
import math
import os
import timeit
import uuid
import warnings
from typing import Optional

import xlsxwriter
from alive_progress import alive_bar
from colorama import init as colorama_init_, Fore
from dotenv import load_dotenv
from pandas import DataFrame
from pymongo import MongoClient

from mongov.constants import *
from mongov.utils import to_str_datetime, serialize_obj, csv_concurrent_, excel_concurrent_, concurrent_

load_dotenv(verbose=True)
colorama_init_(autoreset=True)


def check_folder_path(folder_path):
    if folder_path is None:
        _ = '.'
    elif not os.path.exists(folder_path):
        os.makedirs(folder_path)
        _ = folder_path
    else:
        _ = folder_path
    return _


class MongoEngine:

    def __init__(
            self,
            host: Optional[str] = None,
            port: Optional[int] = None,
            username: Optional[str] = None,
            password: Optional[str] = None,
            database: Optional[str] = None,
            collection: Optional[str] = None,
            conn_timeout: Optional[int] = 30,
            conn_retries: Optional[int] = 5
    ):
        self.host = MONGO_HOST if host is None else host
        self.port = MONGO_PORT if port is None else port
        self.username = username
        self.password = password
        self.database = MONGO_DATABASE if database is None else database
        self.collection = MONGO_COLLECTION if collection is None else collection
        self.conn_timeout = MONGO_CONN_TIMEOUT if conn_timeout is None else conn_timeout
        self.conn_retries = MONGO_CONN_RETRIES if conn_retries is None else conn_retries
        self.mongo_core_ = MongoClient(
            host=self.host,
            port=self.port,
            username=self.username,
            password=self.password,
            maxPoolSize=200
        )
        self.db_ = self.mongo_core_[self.database]
        self.collection_names = self.db_.list_collection_names()
        if self.collection:
            self.collection_ = self.db_[self.collection]
        else:
            self.collection_ = None

    def save_csv_(self, pg, block_size_, collection_name, folder_path_):
        print("线程启动 ...")
        doc_list_ = self.collection_.find({}, {"_id": 0}, batch_size=block_size_).skip(pg * block_size_).limit(
            block_size_)
        filename = f'{collection_name}_{pg}.csv'
        # filename = f'{collection_name}_{str(uuid.uuid4())}.csv'
        with codecs.open(f'{folder_path_}/{filename}', 'w', encoding=PANDAS_ENCODING) as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(list(dict(doc_list_[0]).keys()))
            writer.writerows([list(dict(data_).values()) for data_ in doc_list_])
        return f'{Fore.GREEN} → {folder_path_}/{filename} is ok'

    def to_csv(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1,
               is_block: bool = False, block_size: int = 1000):
        """
        :param query: dict type → Invalid when exporting multiple tables
        """
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, bool):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)

        if self.collection_:
            stats_ = self.db_.command('collstats', self.collection)
            print(f'命名空间: {stats_.get("ns")}, '
                  f'内存总大小: {round(stats_.get("size") / 1024, 2)} KB, '
                  f'存储大小: {round(stats_.get("storageSize") / 1024, 2)} KB, '
                  f'对象平均大小: {round(stats_.get("avgObjSize") / 1024, 2)} KB, '
                  f'文档数: {stats_.get("count")}',
                  )

            if filename is None:
                filename = f'{self.collection}_{to_str_datetime()}.csv'
            start_ = timeit.default_timer()

            if is_block:
                count_ = self.collection_.count_documents(query)
                block_count_ = math.ceil(count_ / block_size)
                print('线程数: ', block_count_)
                # start_ = timeit.default_timer()
                csv_concurrent_(self.save_csv_, self.collection, block_count_, block_size, folder_path_)
                result_ = ECHO_INFO.format(Fore.GREEN, self.collection, folder_path_)
                stop_ = timeit.default_timer()
                print(f'Time: {stop_ - start_}')
            else:
                doc_list = self.collection_.find(query, {'_id': 0}).limit(limit) \
                    if limit != -1 else self.collection_.find(query, {'_id': 0})

                """
                todo polars
                [{"a":111,"b":222},{"a":111,"b":222}] -> [{"a":[111,111],"b":[222,222]}]
                doc_list_ = [list(doc.values()) for doc in doc_list]
                columns_ = list(doc_list[0].keys())
                data = pl.DataFrame(doc_list_,columns=columns_)
                data.to_csv(file=f'{folder_path_}/{filename}', has_header=True)

                todo pandas
                df_ = DataFrame(data=doc_list)
                df_.to_csv(path_or_buf=f'{folder_path_}/{filename}', index=False, encoding=PANDAS_ENCODING)
                """

                title_ = f'{Fore.GREEN} {self.collection} → {folder_path_}/{filename}'
                count_ = self.collection_.count_documents(query)
                with alive_bar(count_, title=title_, bar="blocks") as bar:
                    with codecs.open(f'{folder_path_}/{filename}', 'w', 'utf-8') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(list(dict(doc_list[0]).keys()))
                        # writer.writerows([list(dict(data_).values()) for data_ in doc_list])
                        for data_ in doc_list:
                            writer.writerow(list(dict(data_).values()))
                            bar()

                result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
                stop_ = timeit.default_timer()
                print(f'Time: {stop_ - start_}')
            return result_
        else:
            warnings.warn('No collection specified, All collections will be exported.', DeprecationWarning)
            self.to_csv_s_(folder_path_)
            result_ = ECHO_INFO.format(Fore.GREEN, self.database, folder_path_)
            return result_

    def save_excel_(self, f_: xlsxwriter.Workbook, pg: int, block_size_: int, collection_name: str, folder_path_: str):
        print("线程启动 ...")
        if f_:
            work_sheet_ = f_.add_worksheet(f'Sheet{pg+1}')
            doc_objs_ = self.collection_.find({}, {"_id": 0}, batch_size=block_size_).skip(pg * block_size_).limit(
                block_size_)
            # print(f"A{i + 1}", list(dict(doc).values()))
            for i, doc in enumerate(doc_objs_):
                work_sheet_.write_row(f"A{i + 1}", [
                    '重庆江景登高看才显得壮观，但长江索道排队很长，南山一棵树人多拥挤。最后我选择了鹅岭公园看江景，公园里瞰胜楼五块钱的门票爬到顶层，就能俯瞰重庆雾气缭绕的江景。一边长江一边嘉陵江，能看到颜色的差别，遗憾之处就是雾气笼罩着江水，有点模糊，但人就是要接受一些美中不足。鹅岭公园本身也很漂亮，里面有一些园林设计，植被茂盛，很有南方的感觉。',
                    '重庆', '8e712402-108e-4067-9daf-19d9a60345db', '[]', '穷游', '2020-12-12', '无分类', 131246, '3', '2021-10-12 16:54:40', '其实我是岛酱', 1, '2021-10-12'])
            return f'{Fore.GREEN} → {work_sheet_.name} is ok'
        else:
            filename = f'{collection_name}_{pg}.xlsx'
            xf_ = xlsxwriter.Workbook(filename=f'{folder_path_}/{filename}')
            work_sheet_ = xf_.add_worksheet(f'Sheet{pg + 1}')
            doc_list_ = self.collection_.find({}, {"_id": 0}, batch_size=block_size_).skip(pg * block_size_).limit(
                block_size_)
            for i, doc in enumerate(doc_list_):
                work_sheet_.write_row(f"A{i + 1}", [
                    '重庆江景登高看才显得壮观，但长江索道排队很长，南山一棵树人多拥挤。最后我选择了鹅岭公园看江景，公园里瞰胜楼五块钱的门票爬到顶层，就能俯瞰重庆雾气缭绕的江景。一边长江一边嘉陵江，能看到颜色的差别，遗憾之处就是雾气笼罩着江水，有点模糊，但人就是要接受一些美中不足。鹅岭公园本身也很漂亮，里面有一些园林设计，植被茂盛，很有南方的感觉。',
                    '重庆', '8e712402-108e-4067-9daf-19d9a60345db', '[]', '2021-10-12'])
            xf_.close()
            return f'{Fore.GREEN} → {xf_.filename} is ok'

    def to_excel(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1,
                 is_block: bool = False, block_size: int = 1000, mode: str = 'xlsx'):
        """
        :param query:
        :param folder_path:
        :param filename:
        :param _id:
        :param limit:
        :param is_block:
        :param block_size:
        :param mode: ['sheet' 'xlsx']
        :return:
        """
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, bool):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)

        if self.collection_:
            if filename is None:
                filename = f'{self.collection}_.xlsx'
            start_ = timeit.default_timer()
            doc_objs_ = self.collection_.find(query, {'_id': 0}).limit(
                limit) if limit != -1 else self.collection_.find(query, {'_id': 0})
            if is_block:
                count_ = self.collection_.count_documents(query)
                block_count_ = math.ceil(count_ / block_size)
                print('线程数: ', block_count_)
                if mode not in ['xlsx', 'sheet']:
                    raise ValueError("mode must be specified as xlsx or sheet")
                # 导出为多 xlsx
                if mode == 'xlsx':
                    excel_concurrent_(self.save_excel_, None, self.collection, block_count_, block_size, folder_path_)
                    result_ = ECHO_INFO.format(Fore.GREEN, self.collection, folder_path_)
                    stop_ = timeit.default_timer()
                    print(f'Time: {stop_ - start_}')
                # 导出为多 sheet
                else:
                    f_ = xlsxwriter.Workbook(filename=f'{folder_path_}/{filename}')
                    excel_concurrent_(self.save_excel_, f_, self.collection, block_count_, block_size, folder_path_)
                    f_.close()
                    result_ = ECHO_INFO.format(Fore.GREEN, self.collection, folder_path_)
                    stop_ = timeit.default_timer()
                    print(f'Time: {stop_ - start_}')
            else:
                """
                todo xlwings
                import xlwings as xw
                app = xw.App(visible=False, add_book=False)
                wb = app.books.open(f'{folder_path_}/{filename}') # 打开Excel文件
                sheet = wb.sheets[0]
                for index, doc in enumerate(doc_objs_):
                    sheet.range(f'A{index+1}').value = list(dict(doc).values())
                wb.save()
                wb.close()

                todo openpyxl
                from openpyxl import Workbook
                wb = Workbook(write_only=True)
                ws = wb.create_sheet(title="hi")
                for doc in doc_objs_:
                    ws.append(list(dict(doc).values()))
                wb.save(filename)
                wb.close()
                """

                f = xlsxwriter.Workbook(filename=f'{folder_path_}/{filename}')  # 创建excel文件
                worksheet1 = f.add_worksheet('操作日志')  # 括号内为工作表表名
                title_ = f'{Fore.GREEN} {self.collection} → {folder_path_}/{filename}'
                count_ = self.collection_.count_documents(query)
                with alive_bar(count_, title=title_, bar="blocks") as bar:
                    for i, doc in enumerate(doc_objs_):
                        worksheet1.write_row(f"A{i + 1}", [
                            '重庆江景登高看才显得壮观，但长江索道排队很长，南山一棵树人多拥挤。最后我选择了鹅岭公园看江景，公园里瞰胜楼五块钱的门票爬到顶层，就能俯瞰重庆雾气缭绕的江景。一边长江一边嘉陵江，能看到颜色的差别，遗憾之处就是雾气笼罩着江水，有点模糊，但人就是要接受一些美中不足。鹅岭公园本身也很漂亮，里面有一些园林设计，植被茂盛，很有南方的感觉。',])
                        bar()
                    f.close()

                """
                todo pandas
                doc_list_ = list(doc_objs_)
                data = DataFrame(doc_list_)
                data.to_excel(excel_writer=f'{folder_path_}/{filename}', sheet_name=filename, engine='xlsxwriter', index=False,encoding=PANDAS_ENCODING)
                """

                stop_ = timeit.default_timer()
                print(f'Time: {stop_ - start_}')
                result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
            return result_
        else:
            warnings.warn('No collection specified, All collections will be exported.', DeprecationWarning)
            self.to_excel_s_(folder_path_)
            result_ = ECHO_INFO.format(Fore.GREEN, self.database, folder_path_)
            return result_

    def to_json(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1):
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, bool):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)

        if self.collection_:
            if filename is None:
                filename = f'{self.collection}_{to_str_datetime()}.json'
            doc_objs_ = self.collection_.find(query, {"_id": 0}).limit(
                limit) if limit != -1 else self.collection_.find(query, {"_id": 0})
            doc_list_ = list(doc_objs_)
            data = {'RECORDS': doc_list_}
            with open(f'{folder_path_}/{filename}', 'w', encoding="utf-8") as f:
                f.write(serialize_obj(data))
            result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
            return result_
        else:
            warnings.warn('No collection specified, All collections will be exported.', DeprecationWarning)
            self.to_json_s_(folder_path_)
            result_ = ECHO_INFO.format(Fore.GREEN, self.database, folder_path_)
            return result_

    def to_pickle(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1):
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, int):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)

        if self.collection_:
            if filename is None:
                filename = f'{self.collection}_{to_str_datetime()}.pkl'
            doc_objs_ = self.collection_.find(query, {"_id": 0}).limit(
                limit) if limit != -1 else self.collection_.find(query, {"_id": 0})
            doc_list_ = list(doc_objs_)
            data = DataFrame(doc_list_)
            data.to_pickle(path=f'{folder_path_}/{filename}')
            result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
            return result_

    def to_feather(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1):
        """
        pip[conda] install pyarrow
        """
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, bool):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)
        if self.collection_:
            if filename is None:
                filename = f'{self.collection}_{to_str_datetime()}.feather'
            doc_objs_ = self.collection_.find(query, {"_id": 0}).limit(
                limit) if limit != -1 else self.collection_.find(query, {"_id": 0})
            doc_list_ = list(doc_objs_)
            data = DataFrame(doc_list_)
            data.to_feather(path=f'{folder_path_}/{filename}')
            result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
            return result_

    def to_parquet(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1):
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, bool):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)
        if self.collection_:
            if filename is None:
                filename = f'{self.collection}_{to_str_datetime()}.parquet'
            doc_objs_ = self.collection_.find(query, {"_id": 0}).limit(
                limit) if limit != -1 else self.collection_.find(query, {"_id": 0})
            doc_list_ = list(doc_objs_)
            data = DataFrame(doc_list_)
            data.to_parquet(path=f'{folder_path_}/{filename}')
            result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
            return result_

    def to_hdf5(self, query=None, folder_path: str = None, filename: str = None, _id: bool = False, limit: int = -1):
        """
        pip[conda] install pytables
        """
        if query is None:
            query = {}
        if not isinstance(query, dict):
            raise TypeError('query must be of dict type')
        if not isinstance(limit, int):
            raise TypeError("limit must be an integer type")
        if not isinstance(_id, bool):
            raise TypeError("_id must be an boolean type")
        folder_path_ = check_folder_path(folder_path)
        if self.collection_:
            if filename is None:
                filename = f'{self.collection}_{to_str_datetime()}.h5'
            doc_objs_ = self.collection_.find(query, {"_id": 0}).limit(
                limit) if limit != -1 else self.collection_.find(query, {"_id": 0})
            doc_list_ = list(doc_objs_)
            data = DataFrame(doc_list_)

            data.to_hdf('temp.h5', key='df', mode='w')
            # h5 = pd.HDFStore(f'{folder_path_}/{filename}', 'w', complevel=4, complib='blosc')
            # h5['df_'] = data
            # h5.close()
            result_ = ECHO_INFO.format(Fore.GREEN, self.collection, f'{folder_path_}/{filename}')
            return result_

    def no_collection_to_csv_(self, collection_: str, folder_path: str, _id: bool = False):
        if collection_:
            filename = f'{collection_}_{to_str_datetime()}.csv'
            doc_objs_ = self.db_[collection_].find({}, {"_id": 0})
            data = DataFrame(list(doc_objs_))
            data.to_csv(path_or_buf=f'{folder_path}/{filename}', index=False, encoding=PANDAS_ENCODING)

    def no_collection_to_excel_(self, collection_: str, folder_path: str, _id: bool = False):
        if collection_:
            filename = f'{collection_}_{to_str_datetime()}.xlsx'
            doc_objs_ = self.db_[collection_].find({}, {"_id": 0})
            data = DataFrame(list(doc_objs_))
            data.to_excel(excel_writer=f'{folder_path}/{filename}.xlsx', index=False, encoding=PANDAS_ENCODING)

    def no_collection_to_json_(self, collection_: str, folder_path: str, _id: bool = False):
        if collection_:
            filename = f'{collection_}_{to_str_datetime()}.json'
            doc_objs_ = self.db_[collection_].find({}, {"_id": 0})
            data = {'RECORDS': list(doc_objs_)}
            with open(f'{folder_path}/{filename}', 'w', encoding="utf-8") as f:
                f.write(serialize_obj(data))

    def to_csv_s_(self, folder_path: str):
        concurrent_(self.no_collection_to_csv_, self.collection_names, folder_path)

    def to_excel_s_(self, folder_path: str):
        concurrent_(self.no_collection_to_excel_, self.collection_names, folder_path)

    def to_json_s_(self, folder_path: str):
        concurrent_(self.no_collection_to_json_, self.collection_names, folder_path)


if __name__ == '__main__':
    M = MongoEngine(
        host='192.168.0.141',
        port=27017,
        username='admin',
        password='sanmaoyou_admin_',
        database='sm_admin_test',
        collection='comment'
    )
    # M.to_csv(folder_path="_csv", is_block=True, block_size=20000)
    M.to_excel(folder_path="_excel", is_block=True, block_size=20000, mode='xlsx')
